# Run the following commands to install the libraries and dependencies
# sudo apt install python3-pip
# sudo pip install pandas
# sudo pip install openpyxl
# To run this script
# python3 cable_list_parser_and_db_creator.py
# The script will parse the excel sheet containing all the supported cables.
# The excel sheet needs to be in the same folder as this script and dataframe name should be changed accordingly
# Similarly, the Sheet with cable list inside the excel file usually named "S1 S2 Cable List ..." should match the corresponding dataframe name
# The script will create a header file called "sl_media_data_cable_db.h" which will contain a DB with all supported cables

import pandas as pd
import openpyxl
import re


#create a source file and open it for writing
file1 = open("sl_media_data_cable_db.h", "w") 

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("SlingshotCableCompatibilityMatrix.xlsm")
 
# Define variable to read sheet
#dataframe1 = dataframe.active
dataframe1 = dataframe['S1 S2 Cable List (7.16.25)']

#find the numbe of rows in excel sheet
curr_row = 1
counter = 1
while (1):
    cell_obj = dataframe1.cell(row = curr_row+1, column = 1)
    alpha_num_pn = str(cell_obj.value)
    if alpha_num_pn == "None":
        break
    else:
        counter = counter + 1
    curr_row = curr_row + 1
        

HP_PN = [] #array of PNs which will be sorted later
for row in range(1, counter):
    curr_row = row
    cell_obj = dataframe1.cell(row = curr_row+1, column = 10) #read the length
    length = str(cell_obj.value)
    if length == "HPE Slingshot L1 1x16 Sw Cbl Kit Cray EX" or length == "HPE Slingshot L1 2x16 Sw Cbl Kit Cray EX" or length == "HPE Slingshot L1 1x32 Sw Cbl Kit Cray EX":
        continue #ignore the cable kits
    cell_obj = dataframe1.cell(row = curr_row+1, column = 7) #read the vendor
    cell_value = str(cell_obj.value).strip() #remove whitespace
    if cell_value == "?":
        continue #ignore cables with unknown vendors
    cell_obj = dataframe1.cell(row = curr_row+1, column = 6) #read the type
    cell_value = str(cell_obj.value).strip() #remove whitespace
    if cell_value == "XCVR":
        continue #ignore cables with unknown type
    cell_obj = dataframe1.cell(row = curr_row+1, column = 1)
    #read the alphanumeric PNs
    alpha_num_pn = str(cell_obj.value)
    alpha_num_pn_strip = alpha_num_pn.strip()
    if alpha_num_pn_strip == "?":
        continue #ignore invalid part numbers
    #convert alpha numeric PNs to numeric PNs
    hp_pn_str = re.sub(r'[^0-9]', '', alpha_num_pn)
    hp_pn_int = int(hp_pn_str)
    HP_PN.append(hp_pn_int)

#sort according to PNs
HP_PN.sort()
total_entries = str(len(HP_PN))
print("Total number of valid cables are ", total_entries)

file1.write("/* SPDX-License-Identifier: GPL-2.0 */\n")
file1.write("/* Copyright 2023,2024,2025 Hewlett Packard Enterprise Development LP */\n\n")
file1.write("/* This file is auto-generated through script and should not be modified */\n\n")
file1.write("#ifndef _SL_MEDIA_DATA_CABLE_DB_H_\n")
file1.write("#define _SL_MEDIA_DATA_CABLE_DB_H_\n\n")
file1.write("#include \"sl_media_jack.h\"\n")
file1.write("\nstatic struct sl_media_cable_attr cable_db[] = {\n")

#Algorithm: We take a PN from sorted list and for each PN, we loop through the excel sheet trying to find its associated data.
#Once we find it, we print that data and then delete it from the excel sheet to avoid overwriting (This is important since there are different cables with same PNs)
for i in range(len(HP_PN)):
    curr_row = 1
    while(1):
        cell_obj = dataframe1.cell(row = curr_row+1, column = 10) #read the length
        length = str(cell_obj.value)
        if length == "HPE Slingshot L1 1x16 Sw Cbl Kit Cray EX" or length == "HPE Slingshot L1 2x16 Sw Cbl Kit Cray EX" or length == "HPE Slingshot L1 1x32 Sw Cbl Kit Cray EX":
            curr_row = curr_row + 1
            continue
        cell_obj = dataframe1.cell(row = curr_row+1, column = 7) #read the vendor
        cell_value = str(cell_obj.value).strip() #remove whitespace
        if cell_value == "?":
            curr_row = curr_row + 1
            continue
        cell_obj = dataframe1.cell(row = curr_row+1, column = 6) #read the type
        cell_value = str(cell_obj.value).strip() #remove whitespace
        if cell_value == "XCVR":
            curr_row = curr_row + 1
            continue
        cell_obj = dataframe1.cell(row = curr_row+1, column = 1)
        alpha_num_pn = str(cell_obj.value) #read the alpha numeric PN
        alpha_num_pn_strip = alpha_num_pn.strip()
        if alpha_num_pn_strip == "?":
            curr_row = curr_row + 1
            continue #ignore invalid part numbers
        #convert alpha numeric PNs to numeric PNs
        hp_pn_str = re.sub(r'[^0-9]', '', alpha_num_pn)
        hp_pn_int = int(hp_pn_str)
        if HP_PN[i] == hp_pn_int:
            file1.write("\t{\n")
            file1.write("\t\t.hpe_pn                 = " + str(HP_PN[i]) + ",\n")
            
            cell_obj = dataframe1.cell(row = curr_row+1, column = 7) #read the vendor
            cell_value = str(cell_obj.value).strip() #remove whitespace
            if cell_value == "TE":
                file1.write("\t\t.vendor                 = " + "SL_MEDIA_VENDOR_TE" + ",\n")
            elif cell_value == "Bizlink" or cell_value == "BizLink":
                file1.write("\t\t.vendor                 = " + "SL_MEDIA_VENDOR_BIZLINK" + ",\n")
            elif cell_value == "Hisense":
                file1.write("\t\t.vendor                 = " + "SL_MEDIA_VENDOR_HISENSE" + ",\n")
            elif cell_value == "Coherent (Finisar II-VI)":
                file1.write("\t\t.vendor                 = " + "SL_MEDIA_VENDOR_FINISAR" + ",\n")
            elif cell_value == "Cloud Light":
                file1.write("\t\t.vendor                 = " + "SL_MEDIA_VENDOR_CLOUD_LIGHT" + ",\n")
            elif cell_value == "Molex":
                file1.write("\t\t.vendor                 = " + "SL_MEDIA_VENDOR_MOLEX" + ",\n")
            else:
                file1.write("\t\t.vendor                 = " + "SL_MEDIA_VENDOR_INVALID" + ",\n")

            cell_obj = dataframe1.cell(row = curr_row+1, column = 6) #read the type
            cell_value = str(cell_obj.value).strip() #remove whitespace
            if cell_value == "DAC" or cell_value == "PEC":
                file1.write("\t\t.type                   = " + "SL_MEDIA_TYPE_PEC" + ",\n")
            elif cell_value == "AOC-A" or cell_value == "AOC-D" or cell_value == "AOC":
                file1.write("\t\t.type                   = " + "SL_MEDIA_TYPE_AOC" + ",\n")
            elif cell_value == "POF":
                file1.write("\t\t.type                   = " + "SL_MEDIA_TYPE_POF" + ",\n")
            elif cell_value == "AEC":
                file1.write("\t\t.type                   = " + "SL_MEDIA_TYPE_AEC" + ",\n")
            else:
                file1.write("\t\t.type                   = " + "SL_MEDIA_TYPE_INVALID" + ",\n")

            cell_obj = dataframe1.cell(row = curr_row+1, column = 5) #read the shape
            cell_value = str(cell_obj.value).strip() #remove whitespace
            if cell_value == "Straight":
                file1.write("\t\t.shape                  = " + "SL_MEDIA_SHAPE_STRAIGHT" + ",\n")
            elif cell_value == "Splitter (Y)":
                file1.write("\t\t.shape                  = " + "SL_MEDIA_SHAPE_SPLITTER" + ",\n")
            elif cell_value == "Bifurcated (H)":
                file1.write("\t\t.shape                  = " + "SL_MEDIA_SHAPE_BIFURCATED" + ",\n")
            else:
                file1.write("\t\t.shape                  = " + "SL_MEDIA_SHAPE_INVALID" + ",\n")

            cell_obj = dataframe1.cell(row = curr_row+1, column = 10) #read the length
            length = str(cell_obj.value)
            #num_length = length.rstrip(length[-1]) #remove whitespaces
            whitelist = set('0123456789.')
            num_length = ''.join(filter(whitelist.__contains__, length)) #remove whitespaces and letters
            num_length = float(num_length) #convert string to float
            num_length = num_length * 100 # multiply by 100 to convert meter to cm
            num_length = int(num_length)  #convert float to int
            num_length = str(num_length) #convert int to string
            file1.write("\t\t.length_cm              = " + num_length + ",\n")

            cell_obj = dataframe1.cell(row = curr_row+1, column = 4) #check if SS200 cable
            cell_value = str(cell_obj.value).strip() #remove whitespace
            if cell_value == "?":
                curr_row = curr_row + 1
                continue #ignore invalid vendor
            cell_obj = df.cell(row = curr_row+1, column = 1)
            alpha_num_pn = str(cell_obj.value) #read the alpha numeric PN
            alpha_num_pn_strip = alpha_num_pn.strip()
            if alpha_num_pn_strip == "?":
                curr_row = curr_row + 1
                continue #ignore invalid part numbers
            #convert alpha numeric PNs to numeric PNs
            hp_pn_str = re.sub(r'[^0-9]', '', alpha_num_pn)
            hp_pn_int = int(hp_pn_str)
            if part_nums[i] == hp_pn_int:
                f.write("\t{\n")
                f.write("\t\t.hpe_pn                   = " + str(part_nums[i]) + ",\n")
            
                cell_obj = df.cell(row = curr_row+1, column = 7) #read the vendor
                cell_value = str(cell_obj.value).strip() #remove whitespace
                if cell_value == "TE":
                    f.write("\t\t.vendor                   = " + "SL_MEDIA_VENDOR_TE" + ",\n")
                elif cell_value == "Bizlink" or cell_value == "BizLink":
                    f.write("\t\t.vendor                   = " + "SL_MEDIA_VENDOR_BIZLINK" + ",\n")
                elif cell_value == "Hisense":
                    f.write("\t\t.vendor                   = " + "SL_MEDIA_VENDOR_HISENSE" + ",\n")
                elif cell_value == "Coherent (Finisar II-VI)":
                    f.write("\t\t.vendor                   = " + "SL_MEDIA_VENDOR_FINISAR" + ",\n")
                elif cell_value == "Cloud Light":
                    f.write("\t\t.vendor                   = " + "SL_MEDIA_VENDOR_CLOUD_LIGHT" + ",\n")
                elif cell_value == "Molex":
                    f.write("\t\t.vendor                   = " + "SL_MEDIA_VENDOR_MOLEX" + ",\n")
                else:
                    f.write("\t\t.vendor                   = " + "SL_MEDIA_VENDOR_INVALID" + ",\n")

		# vendor part num
                cell_obj = df.cell(row = curr_row+1, column = 8)
                vendor_str = str(cell_obj.value)
                vendor_pn = vendor_str.replace(" ","")
                if (len(vendor_pn) > 16):
                    f.write("\t\t.vendor_pn                = \"?\",\n")
                else:
                    f.write("\t\t.vendor_pn                = \"" + vendor_pn + "\",\n")

                cell_obj = df.cell(row = curr_row+1, column = 6) #read the type
                cell_value = str(cell_obj.value).strip() #remove whitespace
                curr_type = cell_value
                if cell_value == "DAC" or cell_value == "PEC":
                    f.write("\t\t.type                     = " + "SL_MEDIA_TYPE_PEC" + ",\n")
                elif cell_value == "AOC-A" or cell_value == "AOC-D" or cell_value == "AOC":
                    f.write("\t\t.type                     = " + "SL_MEDIA_TYPE_AOC" + ",\n")
                elif cell_value == "POF":
                    f.write("\t\t.type                     = " + "SL_MEDIA_TYPE_POC" + ",\n")
                elif cell_value == "AEC":
                    f.write("\t\t.type                     = " + "SL_MEDIA_TYPE_AEC" + ",\n")
                elif cell_value == "XCVR":
                    f.write("\t\t.type                     = " + "SL_MEDIA_TYPE_POC" + ",\n")
                else:
                    f.write("\t\t.type                     = " + "SL_MEDIA_TYPE_INVALID" + ",\n")

                cell_obj = df.cell(row = curr_row+1, column = 5) #read the shape
                cell_value = str(cell_obj.value).strip() #remove whitespace
                if cell_value == "Straight":
                    f.write("\t\t.shape                    = " + "SL_MEDIA_SHAPE_STRAIGHT" + ",\n")
                elif cell_value == "Splitter (Y)":
                    f.write("\t\t.shape                    = " + "SL_MEDIA_SHAPE_SPLITTER" + ",\n")
                elif cell_value == "Bifurcated (H)":
                    f.write("\t\t.shape                    = " + "SL_MEDIA_SHAPE_BIFURCATED" + ",\n")
                else:
                    f.write("\t\t.shape                    = " + "SL_MEDIA_SHAPE_INVALID" + ",\n")

                cell_obj = df.cell(row = curr_row+1, column = 10) #read the length
                length = str(cell_obj.value)
                if curr_type == "XCVR":
                     f.write("\t\t.length_cm                = " + "0" + ",\n")
                else:
                    #num_length = length.rstrip(length[-1]) #remove whitespaces
                    whitelist = set('0123456789.')
                    num_length = ''.join(filter(whitelist.__contains__, length)) #remove whitespaces and letters
                    num_length = float(num_length) #convert string to float
                    num_length = num_length * 100 # multiply by 100 to convert meter to cm
                    num_length = int(num_length)  #convert float to int
                    num_length = str(num_length) #convert int to string
                    f.write("\t\t.length_cm                = " + num_length + ",\n")

                cell_obj = df.cell(row = curr_row+1, column = 4) #check if SS200 cable
                cell_value = str(cell_obj.value).strip() #remove whitespace
                if cell_value == "SS200":
                    f.write("\t\t.is_supported_ss200_cable = " + "true" + ",\n")
                else:
                    f.write("\t\t.is_supported_ss200_cable = " + "false" + ",\n")

                cell_obj = df.cell(row = curr_row+1, column = 11) #read the speed
                cell_value = str(cell_obj.value).strip() #remove whitespace
                if cell_value == "200G E" or cell_value == "200Gb":
                    f.write("\t\t.max_speed                = " + "SL_MEDIA_SPEEDS_SUPPORT_CK_200G" + ",\n")
                elif cell_value == "400G E" or cell_value == "400Gb":
                    f.write("\t\t.max_speed                = " + "SL_MEDIA_SPEEDS_SUPPORT_CK_400G" + ",\n")
                elif cell_value == "800Gb":
                    f.write("\t\t.max_speed                = " + "SL_MEDIA_SPEEDS_SUPPORT_CK_800G" + ",\n")
                else:
                    f.write("\t\t.max_speed                = " + "SL_MEDIA_SPEEDS_SUPPORT_INVALID" + ",\n")

                cell_obj = df.cell(row = curr_row+1, column = 6) #read the type
                cell_value = str(cell_obj.value).strip() #remove whitespace
                ##print("cable_type = ", cell_value)
                if cell_value == "DAC" or cell_value == "PEC":
                    f.write("\t\t.serdes_settings.pre1     = " + "0" + ",\n")
                    f.write("\t\t.serdes_settings.pre2     = " + "0" + ",\n")
                    f.write("\t\t.serdes_settings.pre3     = " + "0" + ",\n")
                    f.write("\t\t.serdes_settings.cursor   = " + "100" + ",\n")
                    f.write("\t\t.serdes_settings.post1    = " + "0" + ",\n")
                    f.write("\t\t.serdes_settings.post2    = " + "0" + ",\n")
                elif cell_value == "AEC" or cell_value == "XCVR":
                    f.write("\t\t.serdes_settings.pre1     = " + "-4" + ",\n")
                    f.write("\t\t.serdes_settings.pre2     = " + "0" + ",\n")
                    f.write("\t\t.serdes_settings.pre3     = " + "0" + ",\n")
                    f.write("\t\t.serdes_settings.cursor   = " + "98" + ",\n")
                    f.write("\t\t.serdes_settings.post1    = " + "0" + ",\n")
                    f.write("\t\t.serdes_settings.post2    = " + "0" + ",\n")
                elif cell_value == "AOC":
                    f.write("\t\t.serdes_settings.pre1     = " + "-12" + ",\n")
                    f.write("\t\t.serdes_settings.pre2     = " + "0" + ",\n")
                    f.write("\t\t.serdes_settings.pre3     = " + "0" + ",\n")
                    f.write("\t\t.serdes_settings.cursor   = " + "98" + ",\n")
                    f.write("\t\t.serdes_settings.post1    = " + "-4" + ",\n")
                    f.write("\t\t.serdes_settings.post2    = " + "0" + ",\n")
                else:
                    f.write("\t\t.serdes_settings.pre1     = " + "-20" + ",\n")
                    f.write("\t\t.serdes_settings.pre2     = " + "0" + ",\n")
                    f.write("\t\t.serdes_settings.pre3     = " + "0" + ",\n")
                    f.write("\t\t.serdes_settings.cursor   = " + "116" + ",\n")
                    f.write("\t\t.serdes_settings.post1    = " + "0" + ",\n")
                    f.write("\t\t.serdes_settings.post2    = " + "0" + ",\n")

                # major ver
                cell_obj = df.cell(row = curr_row+1, column = 17)
                cell_value = str(cell_obj.value).strip()
                ##print("straight_major = ", cell_value)
                if cell_value is None or cell_value == "None" or cell_value == "N/A":
                    f.write("\t\t.fw_ver.major             = -1,\n")
                else:
                    f.write("\t\t.fw_ver.major             = " + str(int(cell_value, 16)) + ",\n")

                # major ver
                cell_obj = df.cell(row = curr_row+1, column = 18)
                cell_value = str(cell_obj.value).strip()
                ##print("straight_minor =", cell_value)
                if cell_value is None or cell_value == "None" or cell_value == "N/A":
                    f.write("\t\t.fw_ver.minor             = -1,\n")
                else:
                    f.write("\t\t.fw_ver.minor             = " + str(int(cell_value, 16)) + ",\n")

                # split major ver
                cell_obj = df.cell(row = curr_row+1, column = 19)
                cell_value = str(cell_obj.value).strip()
                ##print("split_major =", cell_value)
                if cell_value is None or cell_value == "None" or cell_value == "N/A":
                    f.write("\t\t.fw_ver.split_major       = -1,\n")
                else:
                    f.write("\t\t.fw_ver.split_major       = " + str(int(cell_value, 16)) + ",\n")

                # split minor ver
                cell_obj = df.cell(row = curr_row+1, column = 20)
                cell_value = str(cell_obj.value).strip()
                ##print("split_minor =", cell_value)
                if cell_value is None or cell_value == "None" or cell_value == "N/A":
                    f.write("\t\t.fw_ver.split_minor       = -1,\n")
                else:
                    f.write("\t\t.fw_ver.split_minor       = " + str(int(cell_value, 16)) + ",\n")

                f.write("\t},\n")

                sheet = wb[sheet_name]
                sheet.delete_rows(curr_row+1, 1)
                break
            
            curr_row = curr_row + 1

wb = load_workbook(infile)

print("Convert:", source1)
df1 = wb[source1]
HP_PN = []
part_nums_get(df1, rows_count(df1), HP_PN)
HP_PN.sort()
print("Total number of valid cables =", str(len(HP_PN)))

print("Convert:", source2)
df2 = wb[source2]
OSFP_HP_PN = []
part_nums_get(df2, rows_count(df2), OSFP_HP_PN)
OSFP_HP_PN.sort()
print("Total number of valid OSFP cables =", str(len(OSFP_HP_PN)))

file1.write("};\n\n")
file1.write("#endif /* _SL_MEDIA_DATA_CABLE_DB_H_ */\n")
file1.close()
